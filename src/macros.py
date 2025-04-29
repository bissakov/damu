import io
import logging
import multiprocessing
import traceback
from datetime import date, datetime
from typing import Dict, List, NamedTuple, Optional, Tuple, Union, cast

import numpy as np
import openpyxl
import openpyxl.styles
import openpyxl.utils
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas._typing import WriteExcelBuffer
from tqdm import tqdm

from src.error import BankNotSupportedError
from src.subsidy import Error, SubsidyContract
from src.utils.db_manager import DatabaseManager
from src.utils.utils import days360, get_column_mapping, save_to_bytes

logger = logging.getLogger("DAMU")


class BankExcelMismatchError(Exception):
    pass


class TotalFalseValueError(Exception):
    pass


class BalanceAfterRepaymentFalseValueError(Exception):
    pass


class Macro(NamedTuple):
    contract_id: str
    macro: Optional[bytes]
    shifted_macro: Optional[bytes]
    df: bytes
    error: Error

    def to_json(self) -> Dict[str, Union[str, bytes, None]]:
        return {
            "id": self.contract_id,
            "macro": self.macro,
            "shifted_macro": self.shifted_macro,
            "df": self.df,
        }

    def save(self, db: DatabaseManager) -> None:
        db.execute(
            """
            INSERT OR REPLACE INTO macros ( id, macro, shifted_macro, df)
            VALUES (:id, :macro, :shifted_macro, :df)
            """,
            self.to_json(),
        )


def clear_range(ws: Worksheet, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def format_style_save(original_df: pd.DataFrame) -> openpyxl.Workbook:
    df_excel_buffer = io.BytesIO()

    df: pd.DataFrame = original_df.copy()

    mapping = get_column_mapping()

    df = cast(pd.DataFrame, df.loc[:, list(mapping.keys())])

    df.rename(columns=mapping, inplace=True)

    with pd.ExcelWriter(cast(WriteExcelBuffer, df_excel_buffer), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    df_excel_buffer.seek(0)

    workbook = openpyxl.load_workbook(df_excel_buffer)
    ws = workbook.worksheets[0]

    font = openpyxl.styles.Font(size=10)
    bold_font = openpyxl.styles.Font(size=10, bold=True)
    header_alignment = openpyxl.styles.Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )
    header_font = openpyxl.styles.Font(bold=True)

    for col_idx in range(1, len(df.columns) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = header_alignment
        header_cell.font = header_font

        header_name = header_cell.value

        match header_name:
            case "Дата погашения основного долга":
                cell_format = "DD.MM.YYYY"
            case "Ставка вознаграждения":
                cell_format = "0.00%"
            case "Кол-во дней" | "Кол-во дней в году":
                cell_format = "0"
            case (
                "Сумма остатка основного долга"
                | "Сумма погашения основного долга"
                | "Сумма вознаграждения, оплачиваемая финансовым агентством"
                | "Сумма вознаграждения, оплачиваемая Получателем"
                | "Итого сумма начисленного вознаграждения"
                | "Сумма рассчитанной субсидии"
                | "Разница между расчетом Банка и Excel"
            ):
                cell_format = "#,##0.00_);-#,##0.00"
            case "Разница между субсидируемой и несубсидируемой частями":
                cell_format = "#,##0.00_);[RED]-#,##0.00"
            case "Соотношение суммы субсидий на итоговую сумму начисленного вознаграждения":
                cell_format = "#,##0.0000_);-#,##0.0000"
            case (
                "Проверка корректности остатка основного долга после произведенного погашения"
                | 'Проверка корректности столбца "Итого начисленного вознаграждения"'
            ):
                cell_format = '"ИСТИНА";;"ЛОЖЬ"'
            case _:
                cell_format = "General"

        for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            cell_count = len(cells)
            for cell_idx, cell in enumerate(cells):
                if cell_idx != 0:
                    cell.number_format = cell_format
                if cell_idx == cell_count - 1:
                    cell.font = bold_font
                else:
                    cell.font = font

    ws.cell(row=ws.max_row, column=1).value = "Всего:"
    ws.cell(row=ws.max_row, column=1).alignment = header_alignment

    clear_range(ws, "G2:O2")

    return workbook


def shift_workbook(source_df: pd.DataFrame, rows: int, cols: int) -> openpyxl.Workbook:
    df: pd.DataFrame = source_df.loc[~source_df["total"]].copy()
    df.drop("total", axis=1, inplace=True)

    df = df[
        [
            "debt_repayment_date",
            "agency_fee_amount",
            "recipient_fee_amount",
            "total_accrued_fee_amount",
            "principal_debt_balance",
        ]
    ]

    df["debt_repayment_date"] = pd.to_datetime(df["debt_repayment_date"]).dt.strftime("%d.%m.%Y")
    df["debt_repayment_date"] = df["debt_repayment_date"].astype(str)

    df.rename(columns=get_column_mapping(), inplace=True)

    df_excel_buffer = io.BytesIO()
    with pd.ExcelWriter(cast(WriteExcelBuffer, df_excel_buffer), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1")
    df_excel_buffer.seek(0)

    source_wb = openpyxl.load_workbook(df_excel_buffer)

    wb = openpyxl.Workbook()

    ws = wb.active
    source_ws = source_wb.active
    clear_range(source_ws, "G2:O2")

    for row in source_ws.iter_rows():
        for cell in row:
            ws.cell(row=cell.row, column=cell.column, value=cell.value)

    last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    range_to_move = f"A1:{last_cell}"

    ws.move_range(range_to_move, rows=rows, cols=cols)

    return wb


def calculate_day_count(debt_repayment_dates: pd.Series, bank: str) -> pd.Series:
    match bank:
        case "АО «Халык-Лизинг»":
            result = debt_repayment_dates.diff().dt.days.astype(float)
            result = pd.Series(np.where((result - 30).abs() < 5, 30.0, result))
            return result
        case (
            'АО "Банк "Bank RBK"'
            | "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)"
            | "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))"
            | "АО «Халык-Лизинг»"
            | 'АО "Лизинг Групп"'
            | 'АО "Нурбанк"'
        ):
            df = pd.DataFrame(
                {
                    "debt_repayment_date": debt_repayment_dates,
                    "prev_date": debt_repayment_dates.shift(),
                }
            )
            day_count = df.apply(
                lambda row: days360(row["prev_date"].date(), row["debt_repayment_date"].date())
                if pd.notnull(row["prev_date"])
                else None,
                axis=1,
            )
            return day_count.astype(float)
        case _:
            return debt_repayment_dates.diff().dt.days.astype(float)


def diff(
    date1: Union[date, datetime, pd.Timestamp],
    date2: Union[date, datetime, pd.Timestamp],
    bank: str,
) -> int:
    match bank:
        case (
            'АО "Банк "Bank RBK"'
            | "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)"
            | "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))"
            | "АО «Халык-Лизинг»"
            | 'АО "Лизинг Групп"'
            | 'АО "Нурбанк"'
        ):
            return days360(date1, date2)
        case _:
            return (date1 - date2).days


def calculate_subsidy_sum_diffs(
    df: pd.DataFrame, contract: SubsidyContract
) -> List[Tuple[int, int, int]]:
    subsidy_sum_diffs: List[Tuple[int, int, int]] = []
    if (
        contract.rate_four_year != 0
        and contract.rate_four_year != contract.rate_one_two_three_year
        and contract.start_date_four_year < contract.end_date
    ):
        mask = df["debt_repayment_date"] > contract.start_date_four_year
        idx = cast(int, mask.idxmax())
        subsidy_sum_diffs.append(
            (
                idx,
                (contract.start_date_four_year - df.loc[idx - 1, "debt_repayment_date"]).days,
                (df.loc[idx, "debt_repayment_date"] - contract.start_date_four_year).days,
            )
        )
        df.loc[mask, "rate"] = contract.rate_four_year
        if (
            contract.rate_five_year != 0
            and contract.rate_five_year != contract.rate_four_year
            and contract.start_date_five_year < contract.end_date
        ):
            mask = df["debt_repayment_date"] > contract.start_date_five_year
            idx = mask.idxmax()
            subsidy_sum_diffs.append(
                (
                    idx,
                    (contract.start_date_five_year - df.loc[idx - 1, "debt_repayment_date"]).days,
                    (df.loc[idx, "debt_repayment_date"] - contract.start_date_five_year).days,
                )
            )
            df.loc[mask, "rate"] = contract.rate_five_year
            if (
                contract.rate_six_seven_year != 0
                and contract.rate_six_seven_year != contract.rate_five_year
                and contract.start_date_six_seven_year < contract.end_date
            ):
                mask = df["debt_repayment_date"] > contract.start_date_six_seven_year
                idx = mask.idxmax()
                subsidy_sum_diffs.append(
                    (
                        idx,
                        (
                            contract.start_date_six_seven_year
                            - df.loc[idx - 1, "debt_repayment_date"]
                        ).days,
                        (
                            df.loc[idx, "debt_repayment_date"] - contract.start_date_six_seven_year
                        ).days,
                    )
                )
                df.loc[mask, "rate"] = contract.rate_six_seven_year

    return subsidy_sum_diffs


def calculate_subsidy_sum(df: pd.DataFrame, bank: str) -> None:
    if bank in {
        'АО "Банк ЦентрКредит"',
        'АО "First Heartland Jusan Bank"',
        'АО "Нурбанк"',
        'АО "ForteBank"',
        'АО "Банк "Bank RBK"',
        'АО "Лизинг Групп"',
        'АО "Казахстанская Иджара Компания"',
    }:
        df["subsidy_sum"] = (
            df["principal_debt_balance"].shift(1)
            * (df["rate"] / df["day_year_count"])
            * df["day_count"]
        )
    elif bank in {
        'АО "Евразийский банк"',
        'АО "Народный Банк Казахстана"',
        "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)",
        "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))",
        "АО «Халык-Лизинг»",
    }:
        df["subsidy_sum"] = (
            df["principal_debt_balance"] * (df["rate"] / df["day_year_count"]) * df["day_count"]
        )


def calculate_principal_balance_check(df: pd.DataFrame, bank: str) -> None:
    if bank in {
        'АО "Банк ЦентрКредит"',
        'АО "First Heartland Jusan Bank"',
        'АО "Нурбанк"',
        'АО "ForteBank"',
        'АО "Банк "Bank RBK"',
    }:
        # FIXME =(B37-C37)=B38
        df["principal_balance_check"] = np.isclose(
            df["principal_debt_balance"].shift(1) - df["principal_debt_repayment_amount"],
            df["principal_debt_balance"],
        )

    elif bank in {
        'АО "Евразийский банк"',
        'АО "Народный Банк Казахстана"',
        "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)",
        "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))",
        "АО «Халык-Лизинг»",
    }:
        if "Bereke Bank" in bank or "Народный Банк" in bank:
            df["principal_balance_check"] = np.isclose(
                df["principal_debt_balance"]
                .shift()
                .fillna(0.0)
                .sub(df["principal_debt_repayment_amount"].shift().fillna(0.0)),
                df["principal_debt_balance"],
            )
        else:
            df["principal_balance_check"] = np.where(
                df.index == df.index[-1],
                np.isclose(
                    df["principal_debt_balance"] - df["principal_debt_repayment_amount"],
                    0.0,
                ),
                np.isclose(
                    df["principal_debt_balance"] - df["principal_debt_repayment_amount"],
                    df["principal_debt_balance"].shift(-1),
                ),
            )

    elif bank in {'АО "Лизинг Групп"', 'АО "Казахстанская Иджара Компания"'}:
        # FIXME =(B37-C37)=B38
        df["principal_balance_check"] = np.isclose(
            df["principal_debt_balance"] - df["principal_debt_repayment_amount"],
            df["principal_debt_balance"].shift(-1),
        )


def create_macro(
    contract: SubsidyContract,
) -> Tuple[bytes, bytes, bytes]:
    summary_df = contract.df.loc[contract.df["total"]].copy()
    df = contract.df.loc[~contract.df["total"]].copy()

    df.drop("total", axis=1, inplace=True)
    summary_df.drop("total", axis=1, inplace=True)

    df = df.loc[df["debt_repayment_date"] <= contract.end_date].copy()

    if df["debt_repayment_date"].isin([contract.start_date]).any():
        df = df.loc[df["debt_repayment_date"] >= contract.start_date].copy()
        df.reset_index(inplace=True, drop=True)
    else:
        mask = df["debt_repayment_date"] > contract.start_date
        insert_idx = mask.idxmax()

        if insert_idx == 0:
            df.loc[-1] = {
                "debt_repayment_date": contract.start_date,
                "principal_debt_balance": contract.loan_amount,
            }
            df.index = df.index + 1
            df.sort_index(inplace=True)
            df.reset_index(inplace=True, drop=True)
        else:
            check_idx = insert_idx - 1

            if (
                df.loc[check_idx, "principal_debt_balance"] == contract.loan_amount
                and not df.loc[
                    check_idx,
                    [
                        "principal_debt_repayment_amount",
                        "agency_fee_amount",
                        "recipient_fee_amount",
                        "total_accrued_fee_amount",
                    ],
                ].any()
            ):
                df.loc[check_idx, "debt_repayment_date"] = contract.start_date
                df = df[df["debt_repayment_date"] >= contract.start_date]
                df.reset_index(inplace=True, drop=True)
            else:
                df = df.iloc[insert_idx::].copy()
                df.loc[-1] = {
                    "debt_repayment_date": contract.start_date,
                    "principal_debt_balance": contract.loan_amount,
                }
                df.index = df.index + 1
                df.sort_index(inplace=True)
                df.reset_index(inplace=True, drop=True)

    df["rate"] = contract.rate_one_two_three_year
    df["day_year_count"] = contract.year_count

    subsidy_sum_diffs = calculate_subsidy_sum_diffs(df, contract)
    df["day_count"] = calculate_day_count(df["debt_repayment_date"], contract.bank)
    calculate_subsidy_sum(df, contract.bank)
    calculate_principal_balance_check(df, contract.bank)

    for idx, diff1, diff2 in subsidy_sum_diffs:
        df.loc[idx, "subsidy_sum"] = (
            (
                df["principal_debt_balance"].shift(1)
                * (df["rate"].shift(1) / df["day_year_count"])
                * diff1
            )
            + (df["principal_debt_balance"].shift(1) * (df["rate"] / df["day_year_count"]) * diff2)
        ).iloc[idx]

    df["bank_excel_diff"] = (df["agency_fee_amount"] - df["subsidy_sum"]).round(2) + 0.0

    df["check_total"] = np.isclose(
        df[["agency_fee_amount", "recipient_fee_amount"]].sum(axis=1),
        df["total_accrued_fee_amount"],
    )
    df["ratio"] = df["agency_fee_amount"] / df["total_accrued_fee_amount"]
    df["difference2"] = df["agency_fee_amount"] - df["recipient_fee_amount"]

    summary_df["principal_debt_balance"] = df["principal_debt_balance"].sum()
    summary_df["principal_debt_repayment_amount"] = df["principal_debt_repayment_amount"].sum()
    summary_df["agency_fee_amount"] = df["agency_fee_amount"].sum()
    summary_df["recipient_fee_amount"] = df["recipient_fee_amount"].sum()
    summary_df["total_accrued_fee_amount"] = df["total_accrued_fee_amount"].sum()
    summary_df["day_count"] = np.nan
    summary_df["rate"] = np.nan
    summary_df["day_year_count"] = np.nan
    summary_df["subsidy_sum"] = df["subsidy_sum"].sum()
    summary_df["bank_excel_diff"] = df["bank_excel_diff"].sum()
    summary_df["check_total"] = np.nan
    summary_df["ratio"] = np.nan
    summary_df["difference2"] = np.nan
    summary_df["principal_balance_check"] = np.nan

    combined_df = pd.concat([df, summary_df], ignore_index=True)

    workbook = format_style_save(combined_df)
    shifted_workbook = shift_workbook(source_df=contract.df, rows=9, cols=1)

    macro_bytes = save_to_bytes(lambda bf: workbook.save(bf))
    shifted_workbook = save_to_bytes(lambda bf: shifted_workbook.save(bf))
    df_bytes = save_to_bytes(lambda bf: df.to_parquet(bf, engine="fastparquet"))

    return macro_bytes, shifted_workbook, df_bytes


def validate_macro(df_bytes: bytes) -> None:
    df = pd.read_parquet(io.BytesIO(df_bytes), engine="fastparquet")
    if not all(df.loc[1:, "bank_excel_diff"].abs() <= 0.01):
        raise BankExcelMismatchError(
            f"Грубые расхождения в колонке 'Разница между расчетом Банка и Excel'"
        )
    if not all(df.loc[1:, "check_total"]):
        raise TotalFalseValueError(
            f"Ложные значения в колонке 'Проверка корректности столбца \"Итого начисленного вознаграждения\"'"
        )
    if not all(df.loc[1:, "principal_balance_check"]):
        raise BalanceAfterRepaymentFalseValueError(
            f"Ложные значения в колонке 'Проверка корректности остатка основного долга после произведенного погашени'"
        )


def process_macro(contract_id: str, db: DatabaseManager) -> Macro:
    if multiprocessing.current_process().name != "MainProcess":
        logging.disable(logging.CRITICAL)

    data = db.execute(
        """
            SELECT c.id,
                   c.start_date,
                   c.end_date,
                   c.loan_amount,
                   c.df,
                   c.bank,
                   c.year_count,
                   ir.rate_one_two_three_year,
                   ir.rate_four_year,
                   ir.rate_five_year,
                   ir.rate_six_seven_year,
                   ir.start_date_one_two_three_year,
                   ir.end_date_one_two_three_year,
                   ir.start_date_four_year,
                   ir.end_date_four_year,
                   ir.start_date_five_year,
                   ir.end_date_five_year,
                   ir.start_date_six_seven_year,
                   ir.end_date_six_seven_year
            FROM contracts AS c
            INNER JOIN interest_rates AS ir ON ir.id = c.id
            LEFT JOIN errors AS e ON e.id = c.id
            WHERE
                e.traceback IS NULL AND
                c.id = ?
            """,
        (contract_id,),
    )
    data = data[0]

    contract = SubsidyContract(*data)

    error = Error(contract_id=contract_id)
    macro_bytes, shifted_macro_bytes, df_bytes, err_trc = None, None, None, None
    try:
        if contract.bank not in {
            'АО "Банк ЦентрКредит"',
            'АО "First Heartland Jusan Bank"',
            'АО "Нурбанк"',
            'АО "ForteBank"',
            'АО "Банк "Bank RBK"',
            'АО "Евразийский банк"',
            'АО "Народный Банк Казахстана"',
            "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)",
            "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))",
            "АО «Халык-Лизинг»",
            'АО "Лизинг Групп"',
            'АО "Казахстанская Иджара Компания"',
        }:
            raise BankNotSupportedError(
                f"Банк/лизинг {contract.bank!r} не поддерживается для сверки."
            )

        macro_bytes, shifted_macro_bytes, df_bytes = create_macro(contract=contract)
        validate_macro(df_bytes)
    except (Exception, BaseException) as err:
        logger.error(f"{err!r}")
        error.traceback = f"{err!r}\n{traceback.format_exc()}"
        error.human_readable = str(err)

    macro = Macro(
        contract_id=contract.contract_id,
        macro=macro_bytes,
        shifted_macro=shifted_macro_bytes,
        df=df_bytes,
        error=error,
    )

    return macro


def process_macros(db: DatabaseManager) -> None:
    contracts = [
        SubsidyContract(*row)
        for row in db.execute(
            """
            SELECT c.id,
                   c.start_date,
                   c.end_date,
                   c.loan_amount,
                   c.df,
                   c.bank,
                   c.year_count,
                   ir.rate_one_two_three_year,
                   ir.rate_four_year,
                   ir.rate_five_year,
                   ir.rate_six_seven_year,
                   ir.start_date_one_two_three_year,
                   ir.end_date_one_two_three_year,
                   ir.start_date_four_year,
                   ir.end_date_four_year,
                   ir.start_date_five_year,
                   ir.end_date_five_year,
                   ir.start_date_six_seven_year,
                   ir.end_date_six_seven_year
            FROM contracts AS c
            INNER JOIN interest_rates AS ir ON ir.id = c.id
            LEFT JOIN errors AS e ON e.id = c.id
            WHERE e.traceback IS NULL
            """
        )
    ]

    err_count = 0
    with tqdm(total=len(contracts)) as pbar:
        for contract in contracts:
            macro = process_macro(contract.contract_id, db)
            macro.error.save(db)
            macro.save(db)
            if macro.error.traceback:
                err_count += 1

            pbar.desc = f"{macro.contract_id} - {err_count} errors"
            pbar.update(1)
