import ast
import io
import itertools
import logging
import re
import traceback
from collections import defaultdict
from collections.abc import Callable, Generator
from datetime import datetime
from pathlib import Path
from typing import NamedTuple, cast

import numpy as np
import openpyxl
import openpyxl.styles
import openpyxl.utils
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas._typing import WriteExcelBuffer

from sverka import macro_formulas
from sverka.error import (
    BalanceAfterRepaymentFalseValueError,
    BankExcelMismatchError,
    BankNotSupportedError,
    TotalFalseValueError,
)
from sverka.structures import COLUMN_MAPPING
from sverka.subsidy import Error, SubsidyContract
from utils.db_manager import DatabaseManager

logger = logging.getLogger("DAMU")


if __debug__:
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_rows", None)


def df_to_bytes(df: pd.DataFrame) -> bytes:
    with io.BytesIO() as buffer:
        df.to_parquet(buffer, engine="fastparquet")
        return buffer.getvalue()


def wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    with io.BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()


class Macro(NamedTuple):
    contract_id: str
    macro: bytes | None
    shifted_macro: bytes | None
    df: bytes | None

    error: Error

    def to_json(self) -> dict[str, str | bytes | None]:
        return {
            "id": self.contract_id,
            "macro": self.macro,
            "shifted_macro": self.shifted_macro,
            "df": self.df,
        }

    def save(self, db: DatabaseManager) -> None:
        db.request(
            """
            INSERT OR REPLACE INTO macros ( id, macro, shifted_macro, df)
            VALUES (:id, :macro, :shifted_macro, :df)
            """,
            self.to_json(),
            req_type="execute",
        )


class ValidationResult(NamedTuple):
    bank_excel_diff_err_cnt: int
    check_total_err_cnt: int
    principal_balance_check_err_cnt: int
    error: Exception | None

    def error_count(self) -> int:
        return (
            self.bank_excel_diff_err_cnt
            + self.check_total_err_cnt
            + self.principal_balance_check_err_cnt
        )

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() == other.error_count()

    def __ne__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() != other.error_count()

    def __lt__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() < other.error_count()

    def __le__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() <= other.error_count()

    def __gt__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() > other.error_count()

    def __ge__(self, other: object) -> bool:
        if not isinstance(other, ValidationResult):
            return NotImplemented
        return self.error_count() >= other.error_count()


class MacroContents(NamedTuple):
    macro_bytes: bytes
    shifted_macro_bytes: bytes
    df_bytes: bytes


class Result(NamedTuple):
    file_path: Path
    contents: MacroContents
    validation: ValidationResult
    calc_subsidy_sum: Callable[[pd.DataFrame], pd.DataFrame]
    calc_principal_balance_check: Callable[[pd.DataFrame], pd.DataFrame]
    calc_day_count: Callable[[pd.Series], pd.Series]

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation == other.validation

    def __ne__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation != other.validation

    def __lt__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation < other.validation

    def __le__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation <= other.validation

    def __gt__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation > other.validation

    def __ge__(self, other: object) -> bool:
        if not isinstance(other, Result):
            return NotImplemented
        return self.validation >= other.validation


def clear_range(ws: Worksheet, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def format_style_save(original_df: pd.DataFrame) -> openpyxl.Workbook:
    df_excel_buffer = io.BytesIO()

    df: pd.DataFrame = original_df.copy()

    df = df.loc[:, list(COLUMN_MAPPING.keys())]
    df.rename(columns=COLUMN_MAPPING, inplace=True)

    with pd.ExcelWriter(
        cast(WriteExcelBuffer, df_excel_buffer), engine="openpyxl"
    ) as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    df_excel_buffer.seek(0)

    workbook = openpyxl.load_workbook(df_excel_buffer)
    ws = workbook.worksheets[0]

    font = openpyxl.styles.Font(size=10)
    bold_font = openpyxl.styles.Font(size=10, bold=True)
    header_alignment = openpyxl.styles.Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )
    header_font = openpyxl.styles.Font(bold=True)

    for col_idx in range(1, len(df.columns) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = header_alignment
        header_cell.font = header_font

        match header_cell.value:
            case "Дата погашения основного долга":
                cell_format = "DD.MM.YYYY"
            case "Ставка вознаграждения":
                # cell_format = "0.00%"
                cell_format = ""
            case "Кол-во дней" | "Кол-во дней в году":
                cell_format = "0"
            case (
                "Сумма остатка основного долга"
                | "Сумма погашения основного долга"
                | "Сумма вознаграждения, оплачиваемая финансовым агентством"
                | "Сумма вознаграждения, оплачиваемая Получателем"
                | "Итого сумма начисленного вознаграждения"
                | "Сумма рассчитанной субсидии"
                | "Разница между расчетом Банка и Excel"
            ):
                # cell_format = "#,##0.00_);-#,##0.00"
                cell_format = r"#0\.00;-#0\.00"
            case "Разница между субсидируемой и несубсидируемой частями":
                # cell_format = "#,##0.00_);[RED]-#,##0.00"
                cell_format = r"#0\.00_);[RED]-#0\.00"
            case "Соотношение суммы субсидий на итоговую сумму начисленного вознаграждения":
                cell_format = "#,##0.0_);-#,##0.0"
            case (
                "Проверка корректности остатка основного долга после произведенного погашения"
                | 'Проверка корректности столбца "Итого начисленного вознаграждения"'
            ):
                cell_format = '"ИСТИНА";;"ЛОЖЬ"'
            case _:
                cell_format = "General"

        for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            cell_count = len(cells)
            for cell_idx, cell in enumerate(cells):
                if cell_idx != 0:
                    cell.number_format = cell_format
                if cell_idx == cell_count - 1:
                    cell.font = bold_font
                else:
                    cell.font = font

    ws.cell(row=ws.max_row, column=1).value = "Всего:"
    ws.cell(row=ws.max_row, column=1).alignment = header_alignment

    clear_range(ws, "G2:O2")

    return workbook


def shift_workbook(
    source_df: pd.DataFrame,
    contract: SubsidyContract,
    rows: int,
    cols: int,
    documents_folder: Path,
) -> openpyxl.Workbook:
    if "total" in source_df:
        df: pd.DataFrame = source_df.loc[~source_df["total"]].copy()
        df.drop("total", axis=1, inplace=True)
    else:
        df = source_df.copy()

    df = df[
        [
            "debt_repayment_date",
            "agency_fee_amount",
            "recipient_fee_amount",
            "total_accrued_fee_amount",
            "principal_debt_balance",
        ]
    ]

    if df.loc[len(df) - 1, "principal_debt_balance"] == 0:
        df["principal_debt_balance"] = df["principal_debt_balance"].shift(1)
        df.loc[0, "principal_debt_balance"] = int(contract.loan_amount * 100)

    df = df[
        (df["agency_fee_amount"] != 0.0)
        & (pd.isna(df["agency_fee_amount"]).notna())
    ].copy()
    df = df.reset_index(drop=True)

    # if df.loc[0, "agency_fee_amount"] == 0:
    #     # first_principal_debt_balance = df.loc[0, "principal_debt_balance"]
    #     # if (
    #     #     pd.isna(first_principal_debt_balance)
    #     #     or first_principal_debt_balance == 0
    #     # ):
    #     df = df.iloc[1:]
    #     df = df.reset_index(drop=True)

    # df = df[df["agency_fee_amount"] != 0].copy()
    # df = df.reset_index(drop=True)

    df["agency_fee_amount"] = (df["agency_fee_amount"] / 100).astype(float)
    df["recipient_fee_amount"] = (df["recipient_fee_amount"] / 100).astype(
        float
    )
    df["total_accrued_fee_amount"] = (
        df["total_accrued_fee_amount"] / 100
    ).astype(float)
    df["principal_debt_balance"] = (df["principal_debt_balance"] / 100).astype(
        float
    )

    df["debt_repayment_date"] = pd.to_datetime(
        df["debt_repayment_date"]
    ).dt.strftime("%d.%m.%Y")
    df["debt_repayment_date"] = df["debt_repayment_date"].astype(str)

    df.rename(columns=COLUMN_MAPPING, inplace=True)

    df_excel_buffer = io.BytesIO()
    with pd.ExcelWriter(
        cast(WriteExcelBuffer, df_excel_buffer), engine="openpyxl"
    ) as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1")
    df_excel_buffer.seek(0)

    source_wb = openpyxl.load_workbook(df_excel_buffer)

    wb = openpyxl.Workbook()

    ws = wb.active
    assert ws is not None
    source_ws = source_wb.active
    assert source_ws is not None
    clear_range(source_ws, "G2:O2")

    for row in source_ws.iter_rows():
        for cell in row:
            r = cell.row or 1
            c = cell.column or 1
            v = cell.value or ""
            ws.cell(row=r, column=c, value=v)

    last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    range_to_move = f"A1:{last_cell}"

    ws.move_range(range_to_move, rows=rows, cols=cols)

    wb.save(documents_folder / "macro.xlsx")

    return wb


def calculate_subsidy_sum_diffs(
    df: pd.DataFrame, contract: SubsidyContract
) -> list[tuple[int, int, int]]:
    subsidy_sum_diffs: list[tuple[int, int, int]] = []
    if (
        contract.rate_four_year != 0
        and contract.rate_four_year != contract.rate_one_two_three_year
        and contract.start_date_four_year < contract.end_date
    ):
        mask = df["debt_repayment_date"] > contract.start_date_four_year
        idx = cast(int, mask.idxmax())
        subsidy_sum_diffs.append(
            (
                idx,
                (
                    contract.start_date_four_year
                    - df.loc[idx - 1, "debt_repayment_date"]
                ).days,
                (
                    df.loc[idx, "debt_repayment_date"]
                    - contract.start_date_four_year
                ).days,
            )
        )
        df.loc[mask, "rate"] = contract.rate_four_year
        if (
            contract.rate_five_year != 0
            and contract.rate_five_year != contract.rate_four_year
            and contract.start_date_five_year < contract.end_date
        ):
            mask = df["debt_repayment_date"] > contract.start_date_five_year
            idx = mask.idxmax()
            subsidy_sum_diffs.append(
                (
                    idx,
                    (
                        contract.start_date_five_year
                        - df.loc[idx - 1, "debt_repayment_date"]
                    ).days,
                    (
                        df.loc[idx, "debt_repayment_date"]
                        - contract.start_date_five_year
                    ).days,
                )
            )
            df.loc[mask, "rate"] = contract.rate_five_year
            if (
                contract.rate_six_seven_year != 0
                and contract.rate_six_seven_year != contract.rate_five_year
                and contract.start_date_six_seven_year < contract.end_date
            ):
                mask = (
                    df["debt_repayment_date"]
                    > contract.start_date_six_seven_year
                )
                idx = mask.idxmax()
                subsidy_sum_diffs.append(
                    (
                        idx,
                        (
                            contract.start_date_six_seven_year
                            - df.loc[idx - 1, "debt_repayment_date"]
                        ).days,
                        (
                            df.loc[idx, "debt_repayment_date"]
                            - contract.start_date_six_seven_year
                        ).days,
                    )
                )
                df.loc[mask, "rate"] = contract.rate_six_seven_year

    return subsidy_sum_diffs


def generate_summary(
    original: pd.DataFrame, original_summary: pd.DataFrame | None = None
) -> pd.DataFrame:
    df = original.copy()

    if original_summary is not None:
        summary_df = original_summary.copy()
    else:
        summary_df = pd.DataFrame()

    summary_df["principal_debt_balance"] = df["principal_debt_balance"].sum()
    summary_df["principal_debt_repayment_amount"] = df[
        "principal_debt_repayment_amount"
    ].sum()
    summary_df["agency_fee_amount"] = df["agency_fee_amount"].sum()
    summary_df["recipient_fee_amount"] = df["recipient_fee_amount"].sum()
    summary_df["total_accrued_fee_amount"] = df[
        "total_accrued_fee_amount"
    ].sum()
    summary_df["day_count"] = np.nan
    summary_df["rate"] = np.nan
    summary_df["day_year_count"] = np.nan
    summary_df["subsidy_sum"] = df["subsidy_sum"].sum()
    summary_df["bank_excel_diff"] = df["bank_excel_diff"].sum()
    summary_df["check_total"] = np.nan
    summary_df["ratio"] = np.nan
    summary_df["difference2"] = np.nan
    summary_df["principal_balance_check"] = np.nan

    summary_df = summary_df.dropna(axis=1, how="all")

    return summary_df


def create_macro(
    file_path: Path,
    contract: SubsidyContract,
    calc_subsidy_sum: Callable[[pd.DataFrame], pd.DataFrame],
    calc_principal_balance_check: Callable[[pd.DataFrame], pd.DataFrame],
    calc_day_count: Callable[[pd.Series], pd.Series],
    documents_folder: Path,
) -> MacroContents:
    # summary_df = contract.df.loc[contract.df["total"]].copy()
    df = contract.df.loc[~contract.df["total"]].copy()

    df.drop("total", axis=1, inplace=True)
    # summary_df.drop("total", axis=1, inplace=True)

    df = df.loc[df["debt_repayment_date"] <= contract.end_date].copy()

    if df["debt_repayment_date"].isin([contract.start_date]).any():
        df = df.loc[df["debt_repayment_date"] >= contract.start_date].copy()
        df.reset_index(inplace=True, drop=True)
    else:
        mask = df["debt_repayment_date"] > contract.start_date
        insert_idx = cast(int, mask.idxmax())

        loan_amount = int(contract.loan_amount * 100)

        if insert_idx == 0:
            # df.loc[-1] = {
            #     "debt_repayment_date": contract.start_date,
            #     "principal_debt_balance": contract.loan_amount,
            # }
            # df.index = df.index + 1

            df = pd.concat(
                [
                    pd.DataFrame(
                        [
                            {
                                "debt_repayment_date": contract.start_date,
                                "principal_debt_balance": loan_amount,
                            }
                        ]
                    ),
                    df,
                ],
                ignore_index=True,
            )
            df.sort_index(inplace=True)
            df.reset_index(inplace=True, drop=True)
        else:
            check_idx = insert_idx - 1

            if (
                df.loc[check_idx, "principal_debt_balance"] == loan_amount
                and not df.loc[
                    check_idx,
                    [
                        "principal_debt_repayment_amount",
                        "agency_fee_amount",
                        "recipient_fee_amount",
                        "total_accrued_fee_amount",
                    ],
                ].any()
            ):
                df.loc[check_idx, "debt_repayment_date"] = contract.start_date
                df = df[df["debt_repayment_date"] >= contract.start_date]
                df.reset_index(inplace=True, drop=True)
            else:
                df = df.iloc[insert_idx::].copy()

                df = pd.concat(
                    [
                        pd.DataFrame(
                            [
                                {
                                    "debt_repayment_date": contract.start_date,
                                    "principal_debt_balance": loan_amount,
                                }
                            ]
                        ),
                        df,
                    ],
                    ignore_index=True,
                )
                df.sort_index(inplace=True)

                # df.loc[-1] = {
                #     "debt_repayment_date": contract.start_date,
                #     "principal_debt_balance": loan_amount,
                # }
                # df.index = df.index + 1
                # df.sort_index(inplace=True)
                df.reset_index(inplace=True, drop=True)

    df["rate"] = contract.rate_one_two_three_year
    df["day_year_count"] = contract.year_count or 360

    subsidy_sum_diffs = calculate_subsidy_sum_diffs(df, contract)

    # df["rate"] = (df["rate"] * 10_000).round().astype("Int64")
    df["day_count"] = calc_day_count(df["debt_repayment_date"])
    df = calc_subsidy_sum(df)
    df = calc_principal_balance_check(df)

    df["bank_excel_diff"] = df["agency_fee_amount"] - df["subsidy_sum"]

    if any((df.loc[1:, "bank_excel_diff"].abs() > 2)):
        for idx, diff1, diff2 in subsidy_sum_diffs:
            num1 = (
                df["principal_debt_balance"].shift(1)
                * df["rate"].shift(1)
                * diff1
            )
            den = df["day_year_count"] * 10_000

            num2 = df["principal_debt_balance"].shift(1) * df["rate"] * diff2

            df.loc[idx, "subsidy_sum"] = (
                (((num1 + den // 2) // den) + ((num2 + den // 2) // den))
                .astype("Int64")
                .iloc[idx]
            )

            # df.loc[idx, "subsidy_sum"] = (
            #     (
            #         df["principal_debt_balance"].shift(1)
            #         * (df["rate"].shift(1) / df["day_year_count"])
            #         * diff1
            #     )
            #     + (df["principal_debt_balance"].shift(1) * (df["rate"] / df["day_year_count"]) * diff2)
            # ).iloc[idx]

        df["bank_excel_diff"] = df["agency_fee_amount"] - df["subsidy_sum"]

    df["check_total"] = (
        df.loc[1:, ["agency_fee_amount", "recipient_fee_amount"]].sum(axis=1)
        == df.loc[1:, "total_accrued_fee_amount"]
    )
    assert df["check_total"].all()
    df["ratio"] = (
        df["agency_fee_amount"] / df["total_accrued_fee_amount"]
    ).round(1)
    df["difference2"] = df["agency_fee_amount"] - df["recipient_fee_amount"]

    summary_df = generate_summary(df)
    combined_df = pd.concat([df, summary_df], ignore_index=True)

    workbook = format_style_save(combined_df)
    workbook.save(file_path)

    shifted_workbook = shift_workbook(
        source_df=contract.df,
        contract=contract,
        rows=9,
        cols=1,
        documents_folder=documents_folder,
    )

    macro_bytes = wb_to_bytes(workbook)
    shifted_bytes = wb_to_bytes(shifted_workbook)
    df_bytes = df_to_bytes(df)

    return MacroContents(macro_bytes, shifted_bytes, df_bytes)


def retry_create_macro(
    result: Result,
    documents_folder: Path,
    contract: SubsidyContract,
    raise_exc: bool = True,
) -> MacroContents:
    df = pd.read_parquet(
        io.BytesIO(result.contents.df_bytes), engine="fastparquet"
    )

    if result.validation.bank_excel_diff_err_cnt > 0:
        _df = df.copy()
        mask = (_df["bank_excel_diff"].abs() > 2).fillna(False)

        human_readable = ""
        for idx in mask[mask].index:
            original_day_count = _df.loc[idx, "day_count"]
            for offset in range(-5, 5, 1):
                offset_day_count = original_day_count + offset
                _df.loc[idx, "day_count"] = offset_day_count
                _df = result.calc_subsidy_sum(_df)

                bank_excel_diff = (
                    _df.loc[idx, "agency_fee_amount"]
                    - _df.loc[idx, "subsidy_sum"]
                )
                _df.loc[idx, "bank_excel_diff"] = bank_excel_diff
                if abs(bank_excel_diff) <= 2:
                    df.loc[idx] = _df.loc[idx]
                    break
            else:
                debt_repayment_date = datetime.strftime(
                    cast(datetime, df.at[idx, "debt_repayment_date"]),
                    "%d.%m.%Y",
                )
                bank_excel_diff = round(df.loc[idx, "bank_excel_diff"] / 100, 2)
                human_readable += f"'Разница между расчетом Банка и Excel' на {debt_repayment_date} равна {bank_excel_diff}\n"

        human_readable = human_readable.strip()
        if human_readable and raise_exc:
            raise BankExcelMismatchError(
                f"Расхождения >0.02 тиын в колонке 'Разница между расчетом Банка и Excel'\n{human_readable}"
            )

        workbook = format_style_save(df)
        workbook.save(result.file_path)

    summary_df = generate_summary(df)
    combined_df = pd.concat([df, summary_df], ignore_index=True)

    workbook = format_style_save(combined_df)
    workbook.save(result.file_path)

    shifted_workbook = shift_workbook(
        source_df=df,
        contract=contract,
        rows=9,
        cols=1,
        documents_folder=documents_folder,
    )

    macro_bytes = wb_to_bytes(workbook)
    shifted_bytes = wb_to_bytes(shifted_workbook)
    df_bytes = df_to_bytes(df)

    return MacroContents(macro_bytes, shifted_bytes, df_bytes)


def validate_macro(df_bytes: bytes) -> ValidationResult:
    df = pd.read_parquet(io.BytesIO(df_bytes), engine="fastparquet")

    bank_excel_diff_err_cnt = int(
        (df.loc[1:, "bank_excel_diff"].abs() > 2).sum()
    )
    check_total_err_cnt = int((~df.loc[1:, "check_total"]).sum())
    principal_balance_check_err_cnt = int(
        (df.loc[1:, "principal_balance_check"] == 0).sum()
    )

    error = None
    if bank_excel_diff_err_cnt > 0:
        error = BankExcelMismatchError(
            "Расхождения >0.02 тиын в колонке 'Разница между расчетом Банка и Excel'"
        )
    if check_total_err_cnt > 0:
        error = TotalFalseValueError(
            "Ложные значения в колонке 'Проверка корректности столбца \"Итого начисленного вознаграждения\"'"
        )
    if principal_balance_check_err_cnt > 0:
        error = BalanceAfterRepaymentFalseValueError(
            "Ложные значения в колонке 'Проверка корректности остатка основного долга после произведенного погашения'"
        )

    return ValidationResult(
        bank_excel_diff_err_cnt=bank_excel_diff_err_cnt,
        check_total_err_cnt=check_total_err_cnt,
        principal_balance_check_err_cnt=principal_balance_check_err_cnt,
        error=error,
    )


def list_formulas(path: Path | str) -> list[str]:
    with open(path, "r") as f:
        tree = ast.parse(f.read(), path)
    return [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]


def iformulas() -> Generator[
    tuple[
        Callable[[pd.DataFrame], pd.DataFrame],
        Callable[[pd.DataFrame], pd.DataFrame],
        Callable[[pd.Series], pd.Series],
    ]
]:
    formulas = list_formulas(macro_formulas.__file__)

    groups = defaultdict(list)
    for n in formulas:
        match = re.match(r"(.+?)(\d+)$", n)
        if not match:
            raise ValueError(
                f"Unable to parse the formula function name - {n=!r}"
            )
        prefix = match.group(1)
        groups[prefix].append(n)

    group_lists = list(groups.values())

    for g1, g2, g3 in itertools.combinations(group_lists, 3):
        for f1, f2, f3 in itertools.product(g1, g2, g3):
            yield (
                getattr(macro_formulas, f1),
                getattr(macro_formulas, f2),
                getattr(macro_formulas, f3),
            )


def generate_macro(
    macros_folder: Path, contract: SubsidyContract, documents_folder: Path
) -> Result:
    results: list[Result] = []
    for (
        calc_subsidy_sum,
        calc_principal_balance_check,
        calc_day_count,
    ) in iformulas():
        file_name = f"{calc_subsidy_sum.__name__}_{calc_principal_balance_check.__name__}_{calc_day_count.__name__}.xlsx"
        file_path = macros_folder / file_name
        contents = create_macro(
            file_path=file_path,
            contract=contract,
            calc_subsidy_sum=calc_subsidy_sum,
            calc_principal_balance_check=calc_principal_balance_check,
            calc_day_count=calc_day_count,
            documents_folder=documents_folder,
        )

        validation = validate_macro(contents.df_bytes)

        results.append(
            Result(
                file_path=file_path,
                contents=contents,
                validation=validation,
                calc_subsidy_sum=calc_subsidy_sum,
                calc_principal_balance_check=calc_principal_balance_check,
                calc_day_count=calc_day_count,
            )
        )

    if not results:
        raise ValueError("Formulas not found or combinations not generated")

    lowest = min(results)

    for result in results:
        if result.file_path == lowest.file_path:
            continue
        result.file_path.unlink()

    return lowest


def process_macro(
    contract_id: str,
    db: DatabaseManager,
    macros_folder: Path,
    documents_folder: Path,
    raise_exc: bool = True,
    skip_pretty_macro: bool = False,
) -> Macro:
    data = db.request(
        """
        SELECT c.id,
               c.start_date,
               c.end_date,
               c.loan_amount,
               c.df,
               c.bank,
               c.year_count,
               ir.rate_one_two_three_year,
               ir.rate_four_year,
               ir.rate_five_year,
               ir.rate_six_seven_year,
               ir.start_date_one_two_three_year,
               ir.end_date_one_two_three_year,
               ir.start_date_four_year,
               ir.end_date_four_year,
               ir.start_date_five_year,
               ir.end_date_five_year,
               ir.start_date_six_seven_year,
               ir.end_date_six_seven_year
        FROM contracts AS c
        INNER JOIN interest_rates AS ir ON ir.id = c.id
        LEFT JOIN errors AS e ON e.id = c.id
        WHERE
            e.traceback IS NULL AND
            c.id = ?
        """,
        (contract_id,),
        req_type="fetch_one",
    )

    contract = SubsidyContract(*data)

    error = Error(contract_id=contract_id)
    macro_bytes, shifted_bytes, df_bytes, _ = None, None, None, None
    try:
        if contract.bank not in {
            'АО "Банк ЦентрКредит"',
            'АО "First Heartland Jusan Bank"',
            'АО "Нурбанк"',
            'АО "ForteBank"',
            'АО "Банк "Bank RBK"',
            'АО "Евразийский банк"',
            'АО "Народный Банк Казахстана"',
            "АО «Bereke Bank» (ранее ДБ АО «Сбербанк»)",
            "АО «Bereke Bank» (дочерний банк Lesha Bank LLC (Public))",
            "АО «Халык-Лизинг»",
            'АО "Лизинг Групп"',
            'АО "Казахстанская Иджара Компания"',
            'АО "ForteLeasing" (ФортеЛизинг)',
        }:
            raise BankNotSupportedError(
                f"Банк/лизинг {contract.bank!r} не поддерживается для сверки."
            )

        if skip_pretty_macro:
            shifted_workbook = shift_workbook(
                source_df=contract.df,
                contract=contract,
                rows=9,
                cols=1,
                documents_folder=documents_folder,
            )
            shifted_bytes = wb_to_bytes(shifted_workbook)
        else:
            result = generate_macro(
                macros_folder=macros_folder,
                contract=contract,
                documents_folder=documents_folder,
            )

            contents = result.contents
            if result.validation.error:
                contents = retry_create_macro(
                    result,
                    documents_folder=documents_folder,
                    contract=contract,
                    raise_exc=raise_exc,
                )

            macro_bytes, shifted_bytes, df_bytes = (
                contents.macro_bytes,
                contents.shifted_macro_bytes,
                contents.df_bytes,
            )

    except (Exception, BaseException) as err:
        logger.exception(err)
        logger.error(f"{err!r}")
        error.traceback = f"{err!r}\n{traceback.format_exc()}"
        error.human_readable = str(err)

    macro = Macro(
        contract_id=contract.contract_id,
        macro=macro_bytes,
        shifted_macro=shifted_bytes,
        df=df_bytes,
        error=error,
    )

    return macro
