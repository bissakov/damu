import logging
import shutil
from pathlib import Path
from typing import Any, Union

import openpyxl
import win32com.client as win32

from src.utils.utils import kill_all_processes


class Excel:
    def __init__(self) -> None:
        self.app = win32.Dispatch("Excel.Application")
        self.app.DisplayAlerts = False

    def __enter__(self) -> "Excel":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        try:
            self.app.Quit()
        except (Exception, BaseException) as err:
            logging.exception(err)
            kill_all_processes(proc_name="EXCEL")
        del self.app


class Workbook:
    def __init__(self, excel: Excel, file_path: Path) -> None:
        self.excel = excel
        self.wb = self.excel.app.Workbooks.Open(str(file_path))

    def __enter__(self) -> "Workbook":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        try:
            self.wb.Close()
        except (Exception, BaseException) as err:
            logging.exception(err)
            kill_all_processes(proc_name="EXCEL")

    def save_as(self, file_path: Path, file_format: int) -> None:
        self.wb.SaveAs(str(file_path), FileFormat=file_format)


class Word:
    def __init__(self) -> None:
        self.app = win32.Dispatch("Word.Application")
        self.app.Visible = False
        self.app.DisplayAlerts = False

    def close(self) -> None:
        try:
            self.app.Quit()
        except (Exception, BaseException) as err:
            logging.exception(err)
            kill_all_processes(proc_name="WINWORD")
        del self.app

    def __enter__(self) -> "Word":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()


class WordDocument:
    def __init__(self, word: Word, file_path: Union[str, Path]) -> None:
        if isinstance(file_path, Path):
            file_path = str(file_path)

        self.word = word
        self.word_doc = self.open(file_path)

    def open(self, file_path: Union[str, Path]) -> Any:
        if isinstance(file_path, Path):
            file_path = str(file_path)
        return self.word.app.Documents.Open(file_path)

    def close(self) -> None:
        try:
            self.word_doc.Close()
        except (Exception, BaseException) as err:
            logging.exception(err)
            kill_all_processes(proc_name="WINWORD")

    def save_as(self, file_path: Union[str, Path], file_format: int) -> None:
        if isinstance(file_path, Path):
            file_path = str(file_path)

        self.word_doc.SaveAs(file_path, FileFormat=file_format)

    def __enter__(self) -> "WordDocument":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()


def is_correct_file(file_path: Path, excel: Excel) -> bool:
    copy_file_path = file_path.with_name(f"copy_{file_path.name}")
    shutil.copyfile(src=file_path, dst=copy_file_path)
    xlsx_file_path = file_path.with_suffix(".xlsx")

    if not xlsx_file_path.exists():
        with Workbook(excel, copy_file_path) as workbook:
            workbook.save_as(xlsx_file_path, 51)

    workbook = openpyxl.load_workbook(xlsx_file_path, data_only=True)
    sheet = workbook.active
    copy_file_path.unlink()

    is_correct = any(
        cell.alignment.horizontal for row in sheet.iter_rows(max_row=20) for cell in row
    )

    return is_correct


def is_file_exported(file_path: Path, excel: Excel) -> tuple[str, bool]:
    file_name = file_path.name
    if not file_path.exists():
        message = f"File '{file_name}' does not exist yet..."
        return message, False
    if file_path.stat().st_size == 0:
        message = f"File '{file_name}' is empty yet..."
        return message, False
    try:
        file_path.rename(file_path)
    except OSError:
        message = f"File '{file_name}' is locked yet..."
        return message, False
    if not is_correct_file(file_path=file_path, excel=excel):
        message = f"File '{file_name}' is not yet exported..."
        return message, False

    message = f"File '{file_name}' exists and ready..."
    return message, True


def convert_report(excel: Excel, source: Path, dist: Path) -> None:
    dist.parent.mkdir(exist_ok=True)
    with Workbook(excel=excel, file_path=source) as workbook:
        workbook.save_as(dist, 51)
    source.unlink()
    logging.info(f"Converted {dist}")
